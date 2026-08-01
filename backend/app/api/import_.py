"""批量导入模块"""
import csv
import io
import os
import json
import logging
import uuid
import tempfile
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.product import Product
from app.core.deps import require_editor
from app.core.consistency import ConsistencyEngine
from app.config import settings


# P0-H3: logger definition (was missing, causing NameError)
logger = logging.getLogger(__name__)

router = APIRouter()

# 带TTL的缓存用于存储上传文件的解析结果
class TTLCache:
    """带TTL的缓存，防止内存泄漏"""
    
    def __init__(self, ttl_seconds: int = 3600, max_size: int = 100):
        self.ttl_seconds = ttl_seconds
        self.max_size = max_size
        self.cache = {}
        self.timestamps = {}
    
    def __setitem__(self, key: str, value: any):
        """设置缓存值（支持字典语法）"""
        self.set(key, value)
    
    def __getitem__(self, key: str) -> any:
        """获取缓存值（支持字典语法）"""
        value = self.get(key)
        if value is None:
            raise KeyError(key)
        return value
    
    def __contains__(self, key: str) -> bool:
        """检查键是否存在（支持in操作符）"""
        return self.get(key) is not None
    
    def __delitem__(self, key: str):
        """删除缓存条目（支持del操作符）"""
        self.delete(key)
    
    def set(self, key: str, value: any):
        """设置缓存值"""
        # 检查缓存大小限制
        if len(self.cache) >= self.max_size:
            self._cleanup_expired()
            # 如果清理后仍然超限，删除最旧的条目
            if len(self.cache) >= self.max_size:
                oldest_key = min(self.timestamps.keys(), key=lambda k: self.timestamps[k])
                del self.cache[oldest_key]
                del self.timestamps[oldest_key]
        
        self.cache[key] = value
        self.timestamps[key] = datetime.now()
    
    def get(self, key: str) -> Optional[any]:
        """获取缓存值，检查是否过期"""
        if key not in self.cache:
            return None
        
        # 检查是否过期
        if datetime.now() - self.timestamps[key] > timedelta(seconds=self.ttl_seconds):
            self.delete(key)
            return None
        
        return self.cache[key]
    
    def delete(self, key: str):
        """删除缓存条目"""
        if key in self.cache:
            del self.cache[key]
        if key in self.timestamps:
            del self.timestamps[key]
    
    def _cleanup_expired(self):
        """清理过期条目"""
        now = datetime.now()
        expired_keys = [k for k, v in self.timestamps.items() 
                       if now - v > timedelta(seconds=self.ttl_seconds)]
        for key in expired_keys:
            self.delete(key)
    
    def __contains__(self, key: str):
        """支持 in 操作符"""
        return self.get(key) is not None
    
    def __delitem__(self, key: str):
        """支持 del 操作符"""
        self.delete(key)


# F-11: Redis-backed cross-worker upload cache (falls back to in-memory)
_redis_upload_cache = None
try:
    from app.core.redis import UploadCache
    _redis_upload_cache = UploadCache
except ImportError:
    logging.getLogger(__name__).error("Redis not available, upload cache disabled")

def _cache_set(file_id: str, data: dict) -> None:
    """跨 worker 缓存写入，优先 Redis"""
    _upload_cache[file_id] = data
    if _redis_upload_cache is not None:
        try:
            _redis_upload_cache.set(file_id, data, ttl_seconds=settings.UPLOAD_CACHE_TTL)
        except Exception:
            pass

def _cache_get(file_id: str) -> dict | None:
    """跨 worker 缓存读取，优先 Redis"""
    cached = None
    if _redis_upload_cache is not None:
        try:
            cached = _redis_upload_cache.get(file_id)
        except Exception:
            pass
    if cached is None:
        try:
            cached = _upload_cache.get(file_id)
        except Exception:
            pass
    return cached

def _cache_delete(file_id: str) -> None:
    """跨 worker 缓存删除"""
    try:
        del _upload_cache[file_id]
    except (KeyError, TypeError):
        pass
    if _redis_upload_cache is not None:
        try:
            _redis_upload_cache.delete(file_id)
        except Exception:
            pass
# 创建缓存实例：使用配置中的TTL和最大大小
_upload_cache = TTLCache(
    ttl_seconds=settings.UPLOAD_CACHE_TTL,
    max_size=settings.UPLOAD_CACHE_MAX_SIZE
)

# 安全的文件扩展名白名单
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

def _sanitize_filename(filename: str) -> str:
    """
    清洗文件名，移除路径分隔符和特殊字符，防止路径遍历攻击

    Args:
        filename: 原始文件名

    Returns:
        清洗后的安全文件名
    """
    if not filename:
        return "unnamed_file"

    # 1. 移除路径分隔符（/、\）
    sanitized = re.sub(r'[/\\]', '', filename)

    # 2. 移除其他危险字符（:、*、?、"、<、>、|）
    sanitized = re.sub(r'[:*?"<>|]', '', sanitized)

    # 3. 移除控制字符（ASCII 0-31）
    sanitized = re.sub(r'[\x00-\x1f\x7f]', '', sanitized)

    # 4. 移除点号开头（防止隐藏文件）和连续点号（防止目录遍历）
    sanitized = re.sub(r'^\.{1,}', '', sanitized)
    sanitized = re.sub(r'\.{2,}', '.', sanitized)

    # 5. 移除空格开头和结尾
    sanitized = sanitized.strip()

    # 6. 如果清洗后为空，返回一个默认名称
    if not sanitized:
        sanitized = "unnamed_file"

    return sanitized

def _validate_file_extension(filename: str) -> tuple[bool, str]:
    """
    验证文件扩展名是否在白名单中

    Args:
        filename: 文件名

    Returns:
        (是否有效, 扩展名或错误信息)
    """
    if not filename:
        return False, "No filename provided"

    # 获取扩展名（最后一个点号之后的内容）
    ext = os.path.splitext(filename)[1]

    if not ext:
        return False, "File has no extension"

    # 转换为小写进行比较，但返回原始扩展名
    if ext.lower() not in ALLOWED_EXTENSIONS:
        return False, f"File extension '{ext}' not allowed. Supported: {', '.join(sorted(ALLOWED_EXTENSIONS))}"

    return True, ext

def _validate_file_path(file_path: str) -> bool:
    """
    验证文件路径是否安全，防止路径遍历

    Args:
        file_path: 文件路径

    Returns:
        是否安全
    """
    # 检查路径中是否包含目录遍历模式
    if '..' in file_path:
        return False

    # 检查路径是否包含控制字符
    if re.search(r'[\x00-\x1f\x7f]', file_path):
        return False

    # 检查路径是否试图访问系统敏感目录
    # 标准化路径分隔符，以便跨平台检查
    normalized_path = file_path.replace('\\', '/')
    path_lower = normalized_path.lower()

    # Linux/Unix 敏感目录
    linux_sensitive = ['/etc/', '/var/', '/usr/', '/bin/', '/sbin/', '/proc/', '/sys/']
    for sensitive_dir in linux_sensitive:
        # 检查路径是否包含敏感目录
        if sensitive_dir in path_lower:
            return False

    # Windows 敏感目录 - 但需要排除用户临时目录
    # 注意：Windows临时目录通常在C:\Users\<username>\AppData\Local\Temp
    # 我们需要检查是否是系统敏感目录，但允许用户临时目录
    windows_sensitive = ['c:/windows', 'c:/system32']
    for sensitive_dir in windows_sensitive:
        # 检查路径是否包含敏感目录（不区分大小写）
        if sensitive_dir in path_lower:
            return False

    # 检查是否在临时目录中（允许用户临时目录）
    temp_dir = tempfile.gettempdir().replace('\\', '/')
    if not normalized_path.startswith(temp_dir):
        # 如果不在临时目录，检查是否在用户目录中（但排除系统敏感目录）
        user_home = os.path.expanduser('~').replace('\\', '/')
        if normalized_path.startswith(user_home):
            # 在用户目录中，但需要确保不是系统敏感目录
            # 检查是否在AppData/Local/Temp或类似的临时目录
            if '/appdata/local/temp' not in path_lower:
                return False
        else:
            # 既不在临时目录也不在用户目录，拒绝
            return False

    return True

# 系统支持的字段
SYSTEM_FIELDS = {
    "sku": {"type": "str", "required": True, "label": "SKU"},
    "product_name_zh": {"type": "str", "required": True, "label": "中文名称"},
    "product_name_en": {"type": "str", "required": True, "label": "英文名称"},
    "category": {"type": "str", "required": True, "label": "品类"},
    "brand": {"type": "str", "required": False, "label": "品牌"},
    "description_zh": {"type": "str", "required": False, "label": "中文描述"},
    "description_en": {"type": "str", "required": False, "label": "英文描述"},
    "price": {"type": "float", "required": False, "label": "价格"},
    "currency": {"type": "str", "required": False, "label": "货币"},
    "stock": {"type": "int", "required": False, "label": "库存"},
    "color_zh": {"type": "str", "required": False, "label": "中文颜色"},
    "color_en": {"type": "str", "required": False, "label": "英文颜色"},
    "material_zh": {"type": "str", "required": False, "label": "中文材质"},
    "material_en": {"type": "str", "required": False, "label": "英文材质"},
    "size": {"type": "str", "required": False, "label": "尺寸"},
    "weight": {"type": "float", "required": False, "label": "重量"},
    "weight_unit": {"type": "str", "required": False, "label": "重量单位"},
    "length": {"type": "float", "required": False, "label": "长"},
    "width": {"type": "float", "required": False, "label": "宽"},
    "height": {"type": "float", "required": False, "label": "高"},
    "dimension_unit": {"type": "str", "required": False, "label": "尺寸单位"},
    "origin": {"type": "str", "required": False, "label": "产地"},
    "model_number": {"type": "str", "required": False, "label": "型号"},
}


def _parse_csv(file_path: str) -> tuple[list[str], list[dict]]:
    """解析 CSV 文件，返回 (列名列表, 数据行列表)"""
    rows = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append(dict(row))
    return headers, rows


def _parse_excel(file_path: str) -> tuple[list[str], list[dict]]:
    """解析 Excel 文件，返回 (列名列表, 数据行列表)"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    rows = []
    for row in rows_iter:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(val) if val is not None else ""
        rows.append(row_dict)
    wb.close()
    return headers, rows


@router.post("/upload")
def upload_file(
    file: UploadFile = File(...),
    current_user=Depends(require_editor),
):
    """上传导入文件并解析"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")

    # 清洗文件名，防止路径遍历攻击
    sanitized_filename = _sanitize_filename(file.filename)

    # 验证扩展名
    is_valid, ext_or_error = _validate_file_extension(sanitized_filename)
    if not is_valid:
        raise HTTPException(status_code=400, detail=ext_or_error)

    ext = ext_or_error

    # 生成安全的文件ID和文件路径
    file_id = str(uuid.uuid4())
    temp_dir = tempfile.gettempdir()

    # 使用UUID重命名文件，避免文件名注入
    safe_filename = f"import_{file_id}{ext}"
    file_path = os.path.join(temp_dir, safe_filename)

    # 验证文件路径安全性
    if not _validate_file_path(file_path):
        raise HTTPException(status_code=400, detail="Invalid file path")

    # 读取文件内容并保存
    content = file.file.read()

    # 验证文件大小（可选：限制最大10MB）
    max_size = 10 * 1024 * 1024  # 10MB
    if len(content) > max_size:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB")

    with open(file_path, "wb") as f:
        f.write(content)

    # 解析文件
    try:
        if ext == ".csv":
            headers, rows = _parse_csv(file_path)
        else:
            headers, rows = _parse_excel(file_path)
    except Exception as e:
        # 清洗错误信息，移除内部路径，防止信息泄露
        import traceback
        error_msg = str(e)
        # 移除文件路径信息
        import re
        error_msg = re.sub(r'File "[^"]*", line \d+', 'File "***", line ***', error_msg)
        error_msg = re.sub(r'File "[^"]*"', 'File "***"', error_msg)
        error_msg = re.sub(r'line \d+', 'line ***', error_msg)
        # 移除目录路径
        error_msg = re.sub(r'[A-Za-z]:\\[^\s]+', '[path]', error_msg)
        error_msg = re.sub(r'/[^\s]+', '[path]', error_msg)
        
        logger.error(f"File parse error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {error_msg}")

    if not headers:
        raise HTTPException(status_code=400, detail="File is empty or has no headers")

    # 自动匹配字段名
    auto_mapping = {}
    for header in headers:
        header_lower = header.lower().strip()
        for field_name, field_info in SYSTEM_FIELDS.items():
            if header_lower == field_name or header_lower == field_info["label"].lower():
                auto_mapping[header] = field_name
                break

    # 缓存解析结果
    _cache_set(file_id, {
        "file_path": file_path,
        "filename": sanitized_filename,
        "headers": headers,
        "rows": rows[:100],  # 最多预览 100 行
        "total_rows": len(rows),
        "auto_mapping": auto_mapping,
    })

    return {
        "file_id": file_id,
        "filename": sanitized_filename,
        "total_rows": len(rows),
        "headers": headers,
        "preview_rows": rows[:5],  # 前 5 行预览
        "auto_mapping": auto_mapping,
        "available_fields": {k: v["label"] for k, v in SYSTEM_FIELDS.items()},
    }


@router.post("/preview")
def preview_import(
    file_id: str,
    field_mapping: Optional[dict] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_editor),
):
    """预览导入结果"""
    if _cache_get(file_id) is None:
        raise HTTPException(status_code=404, detail="File not found or expired")

    cached = _cache_get(file_id)
    mapping = field_mapping or cached["auto_mapping"]

    # 检查必填字段是否都映射了
    missing_required = []
    for field_name, field_info in SYSTEM_FIELDS.items():
        if field_info["required"] and field_name not in mapping.values():
            missing_required.append(field_info["label"])

    # 检查 SKU 重复
    sku_set = set()
    sku_duplicates = []
    all_validate_rows = []
    try:
        with open(cached["file_path"], "r", encoding="utf-8-sig") as _vf:
            import csv as _csv_val
            _v_reader = _csv_val.DictReader(_vf)
            all_validate_rows = [r for r in _v_reader if r]
    except Exception:
        all_validate_rows = cached.get("rows", [])
    for row in all_validate_rows:
        sku = None
        for header, field in mapping.items():
            if field == "sku":
                sku = row.get(header, "").strip()
                break
        if sku:
            if sku in sku_set:
                sku_duplicates.append(sku)
            sku_set.add(sku)

    # 检查必填字段缺失
    rows_with_issues = []
    for i, row in enumerate(cached["rows"][:20]):  # 检查前 20 行
        issues = []
        for header, field in mapping.items():
            if SYSTEM_FIELDS.get(field, {}).get("required"):
                val = row.get(header, "").strip()
                if not val:
                    issues.append(f"必填字段 '{SYSTEM_FIELDS[field]['label']}' 为空")
        if issues:
            rows_with_issues.append({"row": i + 1, "issues": issues})

    return {
        "total_rows": cached["total_rows"],
        "mapped_fields": len(mapping),
        "missing_required": missing_required,
        "sku_duplicates": sku_duplicates,
        "rows_with_issues": rows_with_issues,
        "can_proceed": len(missing_required) == 0 and len(sku_duplicates) == 0,
    }


@router.post("/execute")
def execute_import(
    file_id: str,
    field_mapping: dict,
    mode: str = "create",
    db: Session = Depends(get_db),
    current_user=Depends(require_editor),
):
    """执行批量导入"""
    if _cache_get(file_id) is None:
        raise HTTPException(status_code=404, detail="File not found or expired")

    if mode not in ("create", "update"):
        raise HTTPException(status_code=400, detail="Mode must be 'create' or 'update'")

    cached = _cache_get(file_id)

    # F-36: Prefetch all existing products (N+1 fix)
    existing_products: dict = {}
    try:
        all_products = db.query(Product).all()
        existing_products = {p.sku: p for p in all_products if p.sku}
    except Exception:
        pass

    success_count = 0
    skip_count = 0
    error_count = 0
    errors = []

    # 重新解析全部行 (缓存仅保存前100行用于预览)
    all_rows = []
    try:
        with open(cached["file_path"], "r", encoding="utf-8-sig") as _reparse_f:
            import csv as _csv_reparse
            _reparse_reader = _csv_reparse.DictReader(_reparse_f)
            all_rows = [r for r in _reparse_reader if r]
    except Exception:
        all_rows = cached.get("rows", [])
    
    for i, row in enumerate(all_rows):
        try:
            # 提取字段值
            product_data = {}
            for header, field in field_mapping.items():
                if field in SYSTEM_FIELDS:
                    val = row.get(header, "").strip()
                    if val:
                        if SYSTEM_FIELDS[field]["type"] == "float":
                            product_data[field] = float(val)
                        elif SYSTEM_FIELDS[field]["type"] == "int":
                            product_data[field] = int(val)
                        else:
                            product_data[field] = val

            # 必填字段检查
            if not product_data.get("sku") or not product_data.get("product_name_zh"):
                error_count += 1
                errors.append({"row": i + 1, "error": "Missing required fields (sku or product_name_zh)"})
                continue

            if mode == "create":
                # 检查 SKU 是否已存在
                existing = existing_products.get(product_data["sku"])
                if existing:
                    skip_count += 1
                    continue

                product = Product(
                    **product_data,
                    created_by=current_user.id,
                )
                db.add(product)
                success_count += 1

            elif mode == "update":
                # 按 SKU 查找并更新
                existing = existing_products.get(product_data["sku"])
                if existing:
                    for key, value in product_data.items():
                        if key != "sku":
                            setattr(existing, key, value)
                    existing.updated_at = datetime.utcnow()
                    success_count += 1
                else:
                    # SKU 不存在则创建
                    product = Product(
                        **product_data,
                        created_by=current_user.id,
                    )
                    db.add(product)
                    success_count += 1

        except Exception as e:
            error_count += 1
            errors.append({"row": i + 1, "error": str(e)})

    db.commit()

    # 清理缓存
    try:
        os.unlink(cached["file_path"])
    except OSError:
        pass
    _cache_delete(file_id)

    return {
        "success_count": success_count,
        "skip_count": skip_count,
        "error_count": error_count,
        "errors": errors[:50],  # 最多返回 50 条错误
        "mode": mode,
    }
